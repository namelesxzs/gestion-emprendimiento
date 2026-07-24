"""
🚀 Sistema de Acompañamiento Empresarial UIE - Cadena de Valor
Desarrollado con Streamlit + openpyxl + plotly
Ejecutar con: streamlit run app_uie.py
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date
import os
import io

# ─────────────────────────────────────────────
# CONFIGURACIÓN GENERAL
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="UIE - Cadena de Valor",
    page_icon="🌱",
    layout="wide",
    initial_sidebar_state="expanded",
)

EXCEL_PATH = "Plantilla_Proyecto_UIE_Cadena_Valor.xlsx"

ETAPAS = ["Descubrir", "Incubar", "Informar", "Fomentar", "Financiar"]
ETAPA_COLORS = {
    "Descubrir": "#6366f1",
    "Incubar":   "#06b6d4",
    "Informar":  "#10b981",
    "Fomentar":  "#f59e0b",
    "Financiar": "#ef4444",
}

# ─────────────────────────────────────────────
# ESTILOS CSS PERSONALIZADOS
# ─────────────────────────────────────────────
st.markdown("""
<style>
    /* Sidebar */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1e1b4b 0%, #312e81 100%);
    }
    [data-testid="stSidebar"] * { color: white !important; }

    /* Tarjetas KPI */
    .kpi-card {
        background: white;
        border-radius: 12px;
        padding: 20px 24px;
        box-shadow: 0 1px 6px rgba(0,0,0,0.08);
        border-left: 4px solid;
        margin-bottom: 8px;
    }
    .kpi-value { font-size: 2rem; font-weight: 700; margin: 0; }
    .kpi-label { font-size: 0.85rem; color: #6b7280; margin: 0; }

    /* Badge de etapa */
    .badge {
        display: inline-block;
        padding: 2px 10px;
        border-radius: 20px;
        font-size: 0.78rem;
        font-weight: 600;
        color: white;
    }

    /* Historial clínico */
    .hc-card {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 10px;
        padding: 16px;
        margin-bottom: 12px;
        border-left: 4px solid #6366f1;
    }
    .hc-fecha { font-size: 0.78rem; color: #64748b; }
    .hc-etapa { font-weight: 700; color: #6366f1; }

    /* Sección principal */
    .section-title {
        font-size: 1.3rem;
        font-weight: 700;
        color: #1e1b4b;
        margin-bottom: 4px;
    }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# CARGA Y GUARDADO DE DATOS
# ─────────────────────────────────────────────
@st.cache_data(ttl=0)
def cargar_datos(path=EXCEL_PATH):
    """Lee todas las hojas del Excel y devuelve DataFrames."""
    if not os.path.exists(path):
        st.error(f"⚠️ No se encontró el archivo **{path}**. "
                 "Asegúrate de que esté en la misma carpeta que app_uie.py.")
        st.stop()

    emprendedores = pd.read_excel(path, sheet_name="Emprendedores")
    acompanamientos = pd.read_excel(path, sheet_name="Acompanamientos")
    reuniones = pd.read_excel(path, sheet_name="Reuniones")

    # Normalizar fechas
    for col in ["Fecha_Ingreso"]:
        if col in emprendedores.columns:
            emprendedores[col] = pd.to_datetime(emprendedores[col]).dt.date

    for col in ["Fecha"]:
        if col in acompanamientos.columns:
            acompanamientos[col] = pd.to_datetime(acompanamientos[col]).dt.date
        if col in reuniones.columns:
            reuniones[col] = pd.to_datetime(reuniones[col]).dt.date

    return emprendedores, acompanamientos, reuniones


def guardar_hoja(df, sheet_name, path=EXCEL_PATH):
    """Escribe un DataFrame de vuelta al Excel sin tocar las demás hojas."""
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb[sheet_name]
    ws.delete_rows(2, ws.max_row)          # borra datos, deja encabezado

    for row in df.itertuples(index=False):
        ws.append(list(row))

    wb.save(path)
    st.cache_data.clear()


def next_id(df):
    return int(df["ID"].max()) + 1 if not df.empty else 1


# ─────────────────────────────────────────────
# SIDEBAR - NAVEGACIÓN
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🌱 UIE Cadena de Valor")
    st.markdown("---")
    seccion = st.radio(
        "Navegar a:",
        [
            "📊 Dashboard KPIs",
            "👤 Emprendedores",
            "📋 Acompañamientos",
            "📅 Calendario de Reuniones",
            "➕ Registrar Acompañamiento",
            "📤 Exportar Datos",
        ],
        label_visibility="collapsed",
    )
    st.markdown("---")
    st.caption("Sistema desarrollado para la UIE\nSharePoint · Power Apps · Next.js")


# ─────────────────────────────────────────────
# CARGA INICIAL
# ─────────────────────────────────────────────
emprendedores, acompanamientos, reuniones = cargar_datos()


# ═══════════════════════════════════════════════════════════════
# SECCIÓN 1: DASHBOARD KPIs
# ═══════════════════════════════════════════════════════════════
if seccion == "📊 Dashboard KPIs":
    st.markdown('<p class="section-title">📊 Dashboard de Indicadores</p>', unsafe_allow_html=True)
    st.caption("Vista global de todos los emprendedores en acompañamiento")
    st.markdown("---")

    # ── KPIs globales ──────────────────────────────────────────
    total_emp = len(emprendedores)
    total_acomp = len(acompanamientos)
    activos = len(emprendedores[emprendedores["Estado"] == "Activo"])
    prom_avance = acompanamientos["Avance_%"].mean() if "Avance_%" in acompanamientos.columns else 0

    c1, c2, c3, c4 = st.columns(4)
    def kpi_box(col, valor, label, color):
        col.markdown(
            f'<div class="kpi-card" style="border-color:{color}">'
            f'<p class="kpi-value" style="color:{color}">{valor}</p>'
            f'<p class="kpi-label">{label}</p></div>',
            unsafe_allow_html=True,
        )

    kpi_box(c1, total_emp,        "Total emprendedores",       "#6366f1")
    kpi_box(c2, activos,          "Emprendedores activos",     "#10b981")
    kpi_box(c3, total_acomp,      "Acompañamientos realizados","#06b6d4")
    kpi_box(c4, f"{prom_avance:.0f}%", "Avance promedio",      "#f59e0b")

    st.markdown("### ")

    # ── Gráfico: distribución por etapa ────────────────────────
    col_a, col_b = st.columns(2)

    with col_a:
        st.markdown("**Emprendedores por etapa (Cadena de Valor)**")
        dist = emprendedores["Etapa_UIE"].value_counts().reset_index()
        dist.columns = ["Etapa", "Cantidad"]
        dist["Color"] = dist["Etapa"].map(ETAPA_COLORS)
        fig_pie = px.pie(
            dist, values="Cantidad", names="Etapa",
            color="Etapa", color_discrete_map=ETAPA_COLORS,
            hole=0.45,
        )
        fig_pie.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=300)
        st.plotly_chart(fig_pie, use_container_width=True)

    with col_b:
        st.markdown("**Avance promedio por etapa**")
        if "Etapa" in acompanamientos.columns and "Avance_%" in acompanamientos.columns:
            avance_etapa = acompanamientos.groupby("Etapa")["Avance_%"].mean().reset_index()
            avance_etapa.columns = ["Etapa", "Avance Promedio"]
            fig_bar = px.bar(
                avance_etapa, x="Etapa", y="Avance Promedio",
                color="Etapa", color_discrete_map=ETAPA_COLORS,
                text="Avance Promedio",
            )
            fig_bar.update_traces(texttemplate="%{text:.0f}%", textposition="outside")
            fig_bar.update_layout(showlegend=False, margin=dict(t=10, b=10), height=300, yaxis_range=[0, 110])
            st.plotly_chart(fig_bar, use_container_width=True)

    # ── Gráfico: estado de compromisos ──────────────────────────
    st.markdown("**Estado de compromisos**")
    if "Estado" in acompanamientos.columns:
        est_comp = acompanamientos["Estado"].value_counts().reset_index()
        est_comp.columns = ["Estado", "Cantidad"]
        color_map = {"Pendiente": "#f59e0b", "En proceso": "#06b6d4", "Cumplido": "#10b981"}
        fig_comp = px.bar(
            est_comp, x="Cantidad", y="Estado", orientation="h",
            color="Estado", color_discrete_map=color_map, text="Cantidad",
        )
        fig_comp.update_layout(showlegend=False, margin=dict(t=10, b=10), height=220)
        st.plotly_chart(fig_comp, use_container_width=True)

    # ── Reuniones ───────────────────────────────────────────────
    st.markdown("**Estado de reuniones**")
    if "Estado" in reuniones.columns:
        est_reu = reuniones["Estado"].value_counts().reset_index()
        est_reu.columns = ["Estado", "Cantidad"]
        fig_reu = px.pie(est_reu, values="Cantidad", names="Estado", hole=0.4)
        fig_reu.update_layout(margin=dict(t=10, b=10), height=250)
        st.plotly_chart(fig_reu, use_container_width=True)


# ═══════════════════════════════════════════════════════════════
# SECCIÓN 2: EMPRENDEDORES
# ═══════════════════════════════════════════════════════════════
elif seccion == "👤 Emprendedores":
    st.markdown('<p class="section-title">👤 Gestión de Emprendedores</p>', unsafe_allow_html=True)
    st.markdown("---")

    tab1, tab2 = st.tabs(["📋 Listado", "➕ Nuevo emprendedor"])

    # ── Tab: Listado ──────────────────────────────────────────
    with tab1:
        # Filtros
        col_f1, col_f2 = st.columns(2)
        filtro_etapa = col_f1.multiselect("Filtrar por etapa", ETAPAS, default=ETAPAS)
        filtro_estado = col_f2.multiselect("Filtrar por estado", ["Activo", "Graduado", "Inactivo"], default=["Activo", "Graduado"])

        df_filtrado = emprendedores[
            emprendedores["Etapa_UIE"].isin(filtro_etapa) &
            emprendedores["Estado"].isin(filtro_estado)
        ]

        st.dataframe(
            df_filtrado,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Etapa_UIE": st.column_config.SelectboxColumn("Etapa UIE", options=ETAPAS),
                "Avance_%":  st.column_config.ProgressColumn("Avance", format="%d%%", min_value=0, max_value=100),
            }
        )
        st.caption(f"{len(df_filtrado)} emprendedores encontrados")

        # KPIs individuales al seleccionar
        st.markdown("#### 🔍 KPIs individuales")
        nombres = emprendedores["Nombre"].tolist()
        sel = st.selectbox("Seleccionar emprendedor", nombres)
        if sel:
            emp = emprendedores[emprendedores["Nombre"] == sel].iloc[0]
            emp_id = emp["ID"]
            acomp_emp = acompanamientos[acompanamientos["EmprendedorID"] == emp_id]
            reu_emp   = reuniones[reuniones["EmprendedorID"] == emp_id]

            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Etapa actual",        emp.get("Etapa_UIE", "—"))
            k2.metric("Acompañamientos",     len(acomp_emp))
            k3.metric("Reuniones",           len(reu_emp))
            k4.metric("Avance promedio",     f"{acomp_emp['Avance_%'].mean():.0f}%" if not acomp_emp.empty else "—")

            if not acomp_emp.empty:
                st.markdown("**Historial de avance**")
                fig_line = px.line(
                    acomp_emp.sort_values("Fecha"),
                    x="Fecha", y="Avance_%",
                    markers=True, color_discrete_sequence=["#6366f1"]
                )
                fig_line.update_layout(height=220, margin=dict(t=10,b=10))
                st.plotly_chart(fig_line, use_container_width=True)

    # ── Tab: Nuevo emprendedor ────────────────────────────────
    with tab2:
        with st.form("form_emp"):
            c1, c2 = st.columns(2)
            nombre    = c1.text_input("Nombre completo *")
            empren    = c2.text_input("Nombre del emprendimiento *")
            sector    = c1.text_input("Sector")
            etapa     = c2.selectbox("Etapa UIE", ETAPAS)
            estado    = c1.selectbox("Estado", ["Activo", "Graduado", "Inactivo"])
            responsable = c2.text_input("Responsable")
            correo    = c1.text_input("Correo")
            telefono  = c2.text_input("Teléfono")

            submitted = st.form_submit_button("💾 Guardar emprendedor", type="primary")
            if submitted:
                if not nombre or not empren:
                    st.error("Nombre y emprendimiento son obligatorios.")
                else:
                    nuevo = {
                        "ID": next_id(emprendedores),
                        "Nombre": nombre,
                        "Emprendimiento": empren,
                        "Sector": sector,
                        "Etapa_UIE": etapa,
                        "Estado": estado,
                        "Fecha_Ingreso": date.today(),
                        "Responsable": responsable,
                        "Correo": correo,
                        "Telefono": telefono,
                    }
                    df_new = pd.concat([emprendedores, pd.DataFrame([nuevo])], ignore_index=True)
                    guardar_hoja(df_new, "Emprendedores")
                    st.success(f"✅ Emprendedor **{nombre}** registrado correctamente.")
                    st.rerun()


# ═══════════════════════════════════════════════════════════════
# SECCIÓN 3: ACOMPAÑAMIENTOS (Historial Clínico)
# ═══════════════════════════════════════════════════════════════
elif seccion == "📋 Acompañamientos":
    st.markdown('<p class="section-title">📋 Historial de Acompañamientos</p>', unsafe_allow_html=True)
    st.caption("Vista tipo historial clínico por emprendedor")
    st.markdown("---")

    sel_emp = st.selectbox("Seleccionar emprendedor", emprendedores["Nombre"].tolist())
    emp_id  = emprendedores[emprendedores["Nombre"] == sel_emp].iloc[0]["ID"]
    hist    = acompanamientos[acompanamientos["EmprendedorID"] == emp_id].sort_values("Fecha", ascending=False)

    if hist.empty:
        st.info("Este emprendedor no tiene acompañamientos registrados aún.")
    else:
        for _, row in hist.iterrows():
            color = ETAPA_COLORS.get(row.get("Etapa", ""), "#6366f1")
            avance = row.get("Avance_%", 0)
            estado = row.get("Estado", "")
            estado_color = {"Pendiente": "#f59e0b", "En proceso": "#06b6d4", "Cumplido": "#10b981"}.get(estado, "#94a3b8")

            st.markdown(f"""
            <div class="hc-card" style="border-left-color:{color}">
              <div style="display:flex; justify-content:space-between; margin-bottom:8px">
                <span class="hc-etapa" style="color:{color}">{row.get('Etapa','')}</span>
                <span class="hc-fecha">📅 {row.get('Fecha','')}</span>
              </div>
              <table style="width:100%; font-size:0.88rem">
                <tr>
                  <td><b>🔍 Diagnóstico:</b></td>
                  <td>{row.get('Diagnostico','—')}</td>
                  <td><b>💡 Recomendaciones:</b></td>
                  <td>{row.get('Recomendaciones','—')}</td>
                </tr>
                <tr>
                  <td><b>✅ Compromisos:</b></td>
                  <td>{row.get('Compromisos','—')}</td>
                  <td><b>📊 Estado:</b></td>
                  <td><span style="color:{estado_color}; font-weight:600">{estado}</span></td>
                </tr>
              </table>
              <div style="margin-top:10px">
                <b>Avance: {avance}%</b>
                <div style="background:#e2e8f0; border-radius:8px; height:8px; margin-top:4px">
                  <div style="background:{color}; width:{avance}%; height:8px; border-radius:8px"></div>
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# SECCIÓN 4: CALENDARIO DE REUNIONES
# ═══════════════════════════════════════════════════════════════
elif seccion == "📅 Calendario de Reuniones":
    st.markdown('<p class="section-title">📅 Gestión de Reuniones</p>', unsafe_allow_html=True)
    st.markdown("---")

    tab_ver, tab_nueva = st.tabs(["📋 Ver reuniones", "➕ Nueva / Modificar"])

    with tab_ver:
        # Merge con nombres
        df_reu = reuniones.merge(
            emprendedores[["ID", "Nombre"]], left_on="EmprendedorID", right_on="ID", suffixes=("","_emp")
        )
        filtro_est = st.multiselect(
            "Filtrar por estado", ["Programada", "Reagendada", "Cancelada", "Realizada"],
            default=["Programada", "Reagendada"]
        )
        df_reu_f = df_reu[df_reu["Estado"].isin(filtro_est)]

        for _, row in df_reu_f.iterrows():
            estado = row.get("Estado", "")
            col_map = {"Programada":"#10b981","Reagendada":"#f59e0b","Cancelada":"#ef4444","Realizada":"#6366f1"}
            c = col_map.get(estado, "#94a3b8")
            st.markdown(f"""
            <div style="background:white; border-left:4px solid {c}; border-radius:8px;
                        padding:12px 16px; margin-bottom:8px; box-shadow:0 1px 4px rgba(0,0,0,0.07)">
              <b>{row.get('Nombre','')}</b>
              &nbsp;·&nbsp; 📅 {row.get('Fecha','')} {row.get('Hora','')}
              &nbsp;·&nbsp; <span style="color:{c};font-weight:600">{estado}</span>
              <br><small style="color:#64748b">{row.get('Observaciones','')}</small>
            </div>
            """, unsafe_allow_html=True)

    with tab_nueva:
        with st.form("form_reunion"):
            c1, c2 = st.columns(2)
            emp_sel = c1.selectbox("Emprendedor *", emprendedores["Nombre"].tolist())
            fecha   = c2.date_input("Fecha *", min_value=date.today())
            hora    = c1.time_input("Hora", value=datetime.strptime("09:00", "%H:%M").time())
            accion  = c2.selectbox("Acción", ["Crear", "Reagendar", "Cancelar"])
            obs     = st.text_area("Observaciones")

            estado_map = {"Crear": "Programada", "Reagendar": "Reagendada", "Cancelar": "Cancelada"}

            sub_reu = st.form_submit_button("💾 Guardar reunión", type="primary")
            if sub_reu:
                emp_row = emprendedores[emprendedores["Nombre"] == emp_sel].iloc[0]
                nueva_reu = {
                    "ID":             next_id(reuniones),
                    "EmprendedorID":  emp_row["ID"],
                    "Fecha":          fecha,
                    "Hora":           hora.strftime("%H:%M"),
                    "Estado":         estado_map[accion],
                    "Accion":         accion,
                    "Observaciones":  obs,
                }
                df_new_reu = pd.concat([reuniones, pd.DataFrame([nueva_reu])], ignore_index=True)
                guardar_hoja(df_new_reu, "Reuniones")
                st.success(f"✅ Reunión {accion.lower()}da para {emp_sel} el {fecha}.")
                st.rerun()


# ═══════════════════════════════════════════════════════════════
# SECCIÓN 5: REGISTRAR ACOMPAÑAMIENTO
# ═══════════════════════════════════════════════════════════════
elif seccion == "➕ Registrar Acompañamiento":
    st.markdown('<p class="section-title">➕ Nuevo Acompañamiento</p>', unsafe_allow_html=True)
    st.caption("Registro tipo historial clínico")
    st.markdown("---")

    with st.form("form_acomp"):
        c1, c2 = st.columns(2)
        emp_sel  = c1.selectbox("Emprendedor *", emprendedores["Nombre"].tolist())
        fecha_a  = c2.date_input("Fecha *", value=date.today())
        etapa_a  = c1.selectbox("Etapa de la Cadena de Valor *", ETAPAS)
        estado_a = c2.selectbox("Estado del compromiso", ["Pendiente", "En proceso", "Cumplido"])

        diagnostico = st.text_area("🔍 Diagnóstico del emprendimiento *", height=80)
        recomend    = st.text_area("💡 Recomendaciones", height=80)
        compromisos = st.text_area("✅ Compromisos adquiridos", height=80)
        avance      = st.slider("📊 Nivel de avance (%)", 0, 100, 50)

        sub_ac = st.form_submit_button("💾 Registrar acompañamiento", type="primary")
        if sub_ac:
            if not diagnostico:
                st.error("El diagnóstico es obligatorio.")
            else:
                emp_row = emprendedores[emprendedores["Nombre"] == emp_sel].iloc[0]
                nuevo_ac = {
                    "ID":             next_id(acompanamientos),
                    "EmprendedorID":  emp_row["ID"],
                    "Fecha":          fecha_a,
                    "Etapa":          etapa_a,
                    "Diagnostico":    diagnostico,
                    "Recomendaciones":recomend,
                    "Compromisos":    compromisos,
                    "Avance_%":       avance,
                    "Estado":         estado_a,
                }
                df_new_ac = pd.concat([acompanamientos, pd.DataFrame([nuevo_ac])], ignore_index=True)
                guardar_hoja(df_new_ac, "Acompanamientos")

                # Actualizar etapa del emprendedor
                emprendedores.loc[emprendedores["ID"] == emp_row["ID"], "Etapa_UIE"] = etapa_a
                guardar_hoja(emprendedores, "Emprendedores")

                st.success(f"✅ Acompañamiento registrado para **{emp_sel}** en etapa **{etapa_a}**.")
                st.balloons()
                st.rerun()


# ═══════════════════════════════════════════════════════════════
# SECCIÓN 6: EXPORTAR DATOS
# ═══════════════════════════════════════════════════════════════
elif seccion == "📤 Exportar Datos":
    st.markdown('<p class="section-title">📤 Exportar Información</p>', unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("### 📥 Descargar Excel completo (actualizado)")
    with open(EXCEL_PATH, "rb") as f:
        st.download_button(
            "⬇️ Descargar base de datos Excel",
            data=f,
            file_name="UIE_Cadena_Valor_Export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.markdown("### 📊 Exportar reporte CSV por hoja")
    c1, c2, c3 = st.columns(3)

    buf_emp = io.StringIO()
    emprendedores.to_csv(buf_emp, index=False)
    c1.download_button("Emprendedores CSV", buf_emp.getvalue(), "emprendedores.csv", "text/csv")

    buf_ac = io.StringIO()
    acompanamientos.to_csv(buf_ac, index=False)
    c2.download_button("Acompañamientos CSV", buf_ac.getvalue(), "acompanamientos.csv", "text/csv")

    buf_reu = io.StringIO()
    reuniones.to_csv(buf_reu, index=False)
    c3.download_button("Reuniones CSV", buf_reu.getvalue(), "reuniones.csv", "text/csv")

    st.markdown("### 📋 Vista previa de datos")
    tab_a, tab_b, tab_c = st.tabs(["Emprendedores", "Acompañamientos", "Reuniones"])
    with tab_a: st.dataframe(emprendedores, hide_index=True, use_container_width=True)
    with tab_b: st.dataframe(acompanamientos, hide_index=True, use_container_width=True)
    with tab_c: st.dataframe(reuniones, hide_index=True, use_container_width=True)
